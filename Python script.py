import csv
from openpyxl import Workbook, load_workbook


def calculate_latency_and_jitter(filename):
    latencies = [] # for Bus - Bus latency
    Bus_Comm_latencies = [] # for Bus - Comm core latency
    Comm_App_latencies = [] # for Comm - App core latency
    App_Comm_latencies = [] # for App - Comm core latency
    Comm_Bus_latencies = [] # for Comm - Bus core latency
    cumulative_time = 0 # initial time 0
    zz_awaiting_reference = False  # Track if a reference frame was found
    ff_awaiting_reference = False  # Track if a reference frame was found
    awaiting_accumulation = False  # Track if we need to accumulate times
    reference_packet_time = None # initial reference packet time 0
    reference_payload = None # initial reference payload 0
    reference_frozen = False  # Flag to check if reference time is frozen
    internal_reference = False # for internal time
    cumulative_internal_time = 0 # for internal time cumulational
    Bus_Comm_latency = None
    Comm_App_latency = None
    App_Comm_latency = None
    Comm_Bus_latency = None


    with open(filename, 'r') as csvfile:
        reader = csv.reader(csvfile)

        # Skip the header row
        next(reader, None)

        for row in reader:
            packet_number = int(row[0])  # Packet number
            time_since_previous_packet = float(row[1])  # Time since previous displayed packet
            source = row[2]
            destination = row[3]
            payload = row[7]  # Payload column
            internal_info = row[6]

            # Ignore specific patterns in the payload
            if source.startswith("Siemens") and destination.startswith("TexasInstrum") and payload.startswith("8000808080808080"):
                payload = payload[16:24]  # Skip first 8 bytes
            elif source.startswith("TexasInstrum") and destination.startswith("Siemens") and payload.startswith("80008080808080"):
                payload = payload[14:22]  # Skip first 7 bytes
            elif source.startswith("5a:45:52:20:30:31") and destination.startswith("45:54:48:41:4c:59"):
                internal_info = internal_info[25:35]

            # Check the last just one frame before the starting frame
            if source.startswith("TexasInstrum") and destination.startswith("Siemens") and payload == "00000000":
                zz_awaiting_reference = True  # The next Siemens -> Texas frame will be our reference
            elif source.startswith("TexasInstrum") and destination.startswith("Siemens") and payload == "ffffffff":
                ff_awaiting_reference = True  # The next Siemens -> Texas frame will be our reference

            # Check for the starting frame means from this frame it will start to capture
            if zz_awaiting_reference and source.startswith("Siemens") and destination.startswith("TexasInstrum") and payload == "ffffffff":
                if not reference_frozen:
                    reference_packet_time = time_since_previous_packet
                    reference_frozen = True  # Freeze the reference time
                awaiting_accumulation = True  # Start accumulating after the reference frame
                zz_awaiting_reference = False  # Reset reference frame flag
                reference_payload = 'ffffffff'
            elif ff_awaiting_reference and source.startswith("Siemens") and destination.startswith("TexasInstrum") and payload == "00000000":
                if not reference_frozen:
                    reference_packet_time = time_since_previous_packet
                    reference_frozen = True  # Freeze the reference time
                awaiting_accumulation = True  # Start accumulating after the reference frame
                ff_awaiting_reference = False  # Reset reference frame flag
                reference_payload = '00000000'

            # Accumulate time for next four frames
            if awaiting_accumulation:
                cumulative_time += time_since_previous_packet #this cumulative time is after the internal communication.

            # Check for the Internal Starting Frame
            # Capture the time of Bus-Comm
            if internal_info == "0x00000008":
                internal_reference = True
                Bus_Comm_latency = time_since_previous_packet
                Bus_Comm_latencies.append(round((Bus_Comm_latency) * 1_000_000, 3))

            if internal_reference:
                cumulative_internal_time += time_since_previous_packet
                if internal_info == "0x00000004": # Capture the time of Comm-App
                    Comm_App_latency = cumulative_internal_time - Bus_Comm_latency
                    Comm_App_latencies.append(round((Comm_App_latency) * 1_000_000, 3))
                    cumulative_internal_time = 0

                if internal_info == "0x00000002": # Capture the time of App-Comm
                    App_Comm_latency = cumulative_internal_time
                    App_Comm_latencies.append(round(App_Comm_latency * 1_000_000, 3))
                    cumulative_internal_time = 0
                    internal_reference = False



            # Once frames are accumulated, find the next Texas -> Siemens frame and add cumulative time
            if awaiting_accumulation and source.startswith("TexasInstrum") and destination.startswith("Siemens") and (reference_payload == payload):

                # For total Bus-Bus Latency calculation
                latency = cumulative_time - reference_packet_time
                latencies.append(round(latency * 1_000_000, 3))  # Round to 3 decimal places

                # Capture the time of Comm-Bus
                Comm_Bus_latency = cumulative_time - reference_packet_time - Bus_Comm_latency - Comm_App_latency - App_Comm_latency
                Comm_Bus_latencies.append(round(Comm_Bus_latency * 1_000_000, 3))

                # Reset flags and cumulative time for next sequence
                reference_frozen = False
                cumulative_time = 0
                awaiting_accumulation = False
                reference_packet_time = None
                reference_payload = None

    # Calculate average latency
    average_latency = sum(latencies) / len(latencies) if latencies else 0

    # Find minimum and maximum latencies
    min_latency = min(latencies) if latencies else 0
    max_latency = max(latencies) if latencies else 0

    # Calculate jitter in +/- format
    jitter_minus = round(min_latency - average_latency, 3)
    jitter_plus = round(max_latency - average_latency, 3)

    return (average_latency, jitter_minus, jitter_plus, min_latency, max_latency, latencies, Comm_App_latencies, App_Comm_latencies, Bus_Comm_latencies, Comm_Bus_latencies)

def write_to_excel(filename, latencies, Comm_App_latencies, App_Comm_latencies, Bus_Comm_latencies, Comm_Bus_latencies):
    # Path to the Excel file
    excel_file = '.../demo.xlsx' # write data in demo.xlsx file

    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        workbook = Workbook()

    sheet = workbook.active

    # Write data into the columns
    for i, (latency, comm_app_latency, app_comm_latency, bus_comm_latency, comm_bus_latency) in enumerate(zip(latencies, Comm_App_latencies, App_Comm_latencies, Bus_Comm_latencies, Comm_Bus_latencies), start=2):
        sheet.cell(row=i, column=2, value=latency)
        sheet.cell(row=i, column=4, value=bus_comm_latency)
        sheet.cell(row=i, column=6, value=comm_app_latency)
        sheet.cell(row=i, column=8, value=app_comm_latency)
        sheet.cell(row=i, column=10, value=comm_bus_latency)

    # Save the workbook
    workbook.save(excel_file)

# absolute file path
filename = '.../1_cyclic.csv'  # CSV file path
(average_latency, jitter_minus, jitter_plus, min_latency, max_latency, latencies, Comm_App_latencies, App_Comm_latencies, Bus_Comm_latencies, Comm_Bus_latencies)  = calculate_latency_and_jitter(filename)

# Write the results to an Excel sheet
write_to_excel(filename, latencies, Comm_App_latencies, App_Comm_latencies, Bus_Comm_latencies, Comm_Bus_latencies)

print("Latencies (microseconds):", [f"{latency:.3f}" for latency in latencies])
print("Bus_Comm_latencies (microseconds):", [f"{latency:.3f}" for latency in Bus_Comm_latencies])
print("Comm_App_latencies (microseconds):", [f"{latency:.3f}" for latency in Comm_App_latencies])
print("App_Comm_latencies (microseconds):", [f"{latency:.3f}" for latency in App_Comm_latencies])
print("Comm_Bus_latencies (microseconds):", [f"{latency:.3f}" for latency in Comm_Bus_latencies])

print(f"Min Latency (microseconds): {min_latency:.3f}")
print(f"Max Latency (microseconds): {max_latency:.3f}")

# Print average latency and jitter values formatted to 3 decimal places
print(f"Average Latency (microseconds): {average_latency:.3f}")
print(f"Jitter (microseconds): {jitter_minus:.3f} / {jitter_plus:.3f}")
